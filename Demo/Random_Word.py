import openpyxl
import random

wb = openpyxl.load_workbook("chinese_words.xlsx")
ws = wb["chinese_words"]

row = random.randint(1, ws.max_row)

word = ws.cell(row, 1).value
meaning = ws.cell(row, 2).value
pinin = ws.cell(row, 3).value

print(word)
print(pinin)
print(meaning)
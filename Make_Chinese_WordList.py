
#使用するモジュールのインポート
from selenium import webdriver
'''webの要素検索するのに使用'''
from selenium.webdriver.common.by import By

import openpyxl

#webdriverの作成
driver_path = "driver\\chromedriver.exe"
driver = webdriver.Chrome(executable_path=driver_path)

#中国語単語のwebページにアクセス
driver.get("http://chugokugo-script.net/tango/level2/doushi.html")

chinese_words_html = driver.find_elements(By.CLASS_NAME, "divBunruiC")      #中国語
chinese_meaning_html = driver.find_elements(By.CLASS_NAME, "divBunruiN")    #ピンイン
chinese_pinin_html = driver.find_elements(By.CLASS_NAME, "divBunruiP")      #意味

wb = openpyxl.load_workbook("chinese_words.xlsx")
ws = wb["chinese_words"]

for i in range(len(chinese_words_html)):
     ws.cell(i+1,1,value = chinese_words_html[i].text)      #中国語
     ws.cell(i+1,2,value = chinese_pinin_html[i].text)      #ピンイン
     ws.cell(i+1,3,value = chinese_meaning_html[i].text)    #意味
     

wb.save("chinese_words.xlsx")

#Driver終了
driver.quit()







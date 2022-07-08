import requests

import openpyxl
import random


################## Excelから単語取得 ##################
wb = openpyxl.load_workbook("chinese_words.xlsx")
ws = wb["chinese_words"]

row = random.randint(1, ws.max_row)

word = ws.cell(row, 1).value
pinin = ws.cell(row, 3).value
meaning = ws.cell(row, 2).value

################## Line へ送信 ##################
acc_token = '7fykWSy9F1AxsdFQBUkrZhsqqnspUgYOSqBFjSmoBYA'

def send_line(msg):
    url = 'https://notify-api.line.me/api/notify'
    headers = {'Authorization': 'Bearer ' + acc_token}
    payload = {'message': msg}
    requests.post(url, headers=headers, params=payload)

if __name__ == '__main__':
    #メッセージ送信
    send_line('\n'+word+'\n'+pinin+'\n'+meaning)